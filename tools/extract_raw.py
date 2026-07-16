# -*- coding: utf-8 -*-
"""건강보험심사평가원 「병원정보서비스」 엑셀(분기별 배포)에서
병원급 이상(상급종합·종합병원·병원·요양병원·정신병원·치과병원·한방병원)만 추려
tools/hospitals_raw.json 으로 저장.

원본 다운로드: https://opendata.hira.or.kr/op/opc/selectOpenData.do?sno=11925
  (건강보험심사평가원_전국 병의원 및 약국 현황, "1.병원정보서비스(YYYY.M.).xlsx")
같은 zip의 "5.의료기관별상세정보서비스_03_진료과목정보" 에서 소아청소년과 전문의
보유 여부를 매칭해 kidsCare 플래그로 추가.
"""
import glob
import json
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

TOOLS = Path(__file__).resolve().parent
OUT = TOOLS / "hospitals_raw.json"

HOSPITAL_KINDS = {"상급종합", "종합병원", "병원", "요양병원", "정신병원", "치과병원", "한방병원"}


def find_source(pattern):
    matches = glob.glob(pattern)
    if not matches:
        raise SystemExit(f"원본 파일을 찾을 수 없음: {pattern}")
    return matches[0]


def main():
    src_dir = sys.argv[1] if len(sys.argv) > 1 else None
    if not src_dir:
        raise SystemExit("사용법: py extract_raw.py <HIRA 압축해제 폴더>")

    hosp_path = find_source(str(Path(src_dir) / "1.*.xlsx"))
    subj_path = find_source(str(Path(src_dir) / "5.*.xlsx"))

    print("병원정보서비스:", hosp_path)
    wb = openpyxl.load_workbook(hosp_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    idx = {name: i for i, name in enumerate(header)}

    records = []
    for r in rows:
        kind = r[idx["종별코드명"]]
        if kind not in HOSPITAL_KINDS:
            continue
        records.append({
            "key": r[idx["암호화요양기호"]],
            "name": r[idx["요양기관명"]],
            "kind": kind,
            "sido": r[idx["시도코드명"]],
            "sigungu": r[idx["시군구코드명"]],
            "address": r[idx["주소"]],
            "phone": r[idx["전화번호"]],
            "homepage": r[idx["병원홈페이지"]],
            "totalDoctors": r[idx["총의사수"]],
            "lat": r[idx["좌표(Y)"]],
            "lng": r[idx["좌표(X)"]],
        })
    wb.close()
    print("병원급 이상:", len(records))

    print("진료과목정보:", subj_path)
    wb2 = openpyxl.load_workbook(subj_path, read_only=True)
    ws2 = wb2.active
    rows2 = ws2.iter_rows(values_only=True)
    header2 = list(next(rows2))
    idx2 = {name: i for i, name in enumerate(header2)}
    kids_keys = set()
    er_keys = set()
    for r in rows2:
        subj = r[idx2["진료과목코드명"]]
        if subj == "소아청소년과":
            kids_keys.add(r[idx2["암호화요양기호"]])
        if subj == "응급의학과":
            er_keys.add(r[idx2["암호화요양기호"]])
    wb2.close()
    print("소아청소년과 보유:", len(kids_keys), "| 응급의학과 보유:", len(er_keys))

    for rec in records:
        rec["kidsCare"] = rec["key"] in kids_keys
        rec["erCare"] = rec["key"] in er_keys

    OUT.write_text(json.dumps({"records": records}, ensure_ascii=False), encoding="utf-8")
    print("저장:", OUT, "|", OUT.stat().st_size, "bytes")


if __name__ == "__main__":
    main()
